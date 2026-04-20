import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('dvf_data.json', encoding='utf-8') as f:
    all_data = json.load(f)
with open('dvf_maison_1500.json', encoding='utf-8') as f:
    maison_data = json.load(f)

wb = openpyxl.Workbook()
years = [2021, 2022, 2023, 2024, 2025]

hdr_fill  = PatternFill('solid', fgColor='1a3a5c')
hdr2_fill = PatternFill('solid', fgColor='1f3d12')
alt1_fill = PatternFill('solid', fgColor='EAF4FF')
alt2_fill = PatternFill('solid', fgColor='FFFFFF')
alt3_fill = PatternFill('solid', fgColor='F0FAE8')
alt4_fill = PatternFill('solid', fgColor='FAFFF5')
bold_w    = Font(bold=True, color='FFFFFF', size=11)
center    = Alignment(horizontal='center', vertical='center')
thin      = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def cell(ws, r, c, val, fill=None, fnt=None, fmt=None):
    ce = ws.cell(row=r, column=c, value=val)
    if fill: ce.fill = fill
    if fnt:  ce.font = fnt
    else:    ce.font = Font(size=10)
    ce.alignment = center
    ce.border = thin
    if fmt:  ce.number_format = fmt
    return ce

# ==================== FEUILLE 1 : Résumé annuel ====================
ws1 = wb.active
ws1.title = 'Résumé annuel'

ws1.merge_cells('A1:I1')
c = ws1['A1']
c.value = 'DVF — Prix immobiliers autour de Félix Faure, Nantes'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='0d2137')
c.alignment = center
ws1.row_dimensions[1].height = 28

ws1.merge_cells('A2:I2')
c = ws1['A2']
c.value = 'Source : files.data.gouv.fr/geo-dvf | Apparts r=800m | Maisons r=1500m | Centre : 47.2091, -1.5573'
c.font = Font(italic=True, color='888888', size=9)
c.alignment = center
ws1.row_dimensions[2].height = 16
ws1.row_dimensions[3].height = 6

hdrs = ['Année', 'Type', 'Rayon', 'Nb tx', 'Prix moy €/m²', 'Surf. moy. m²', 'Prix total estimé', 'Var €/m²', 'Var %']
for c_i, h in enumerate(hdrs, 1):
    cell(ws1, 4, c_i, h, fill=hdr_fill, fnt=bold_w)
ws1.row_dimensions[4].height = 22

row = 5
prev = {}
for year in years:
    # Appartements 800m
    items_a = all_data[str(year)]['Appartement']
    avg_a = round(sum(i['ppm2'] for i in items_a) / len(items_a)) if items_a else 0
    avg_surf_a = round(sum(i['surface'] for i in items_a) / len(items_a), 1) if items_a else 0
    key_a = 'Appart'
    var_abs_a = round(avg_a - prev[key_a]) if key_a in prev else ''
    var_pct_a = round((avg_a - prev[key_a]) / prev[key_a] * 100, 1) if key_a in prev else ''
    prev[key_a] = avg_a
    fill_a = alt1_fill if row % 2 == 0 else alt2_fill
    vals_a = [year, 'Appartement', '800 m', len(items_a), avg_a, avg_surf_a, round(avg_a * avg_surf_a), var_abs_a, var_pct_a]
    for c_i, v in enumerate(vals_a, 1):
        fmt = '#,##0 [$€]' if c_i in (5, 7) else None
        ce = cell(ws1, row, c_i, v, fill=fill_a, fmt=fmt)
        if c_i == 9 and isinstance(v, float):
            ce.font = Font(color='1a7a1a' if v > 0 else 'cc2200', bold=True, size=10)
    row += 1

    # Maisons 1500m
    items_m = maison_data['r1500'][str(year)]
    avg_m = round(sum(i['ppm2'] for i in items_m) / len(items_m)) if items_m else 0
    avg_surf_m = round(sum(i['surface'] for i in items_m) / len(items_m), 1) if items_m else 0
    key_m = 'Maison'
    var_abs_m = round(avg_m - prev[key_m]) if key_m in prev else ''
    var_pct_m = round((avg_m - prev[key_m]) / prev[key_m] * 100, 1) if key_m in prev else ''
    prev[key_m] = avg_m
    fill_m = alt3_fill if row % 2 == 0 else alt4_fill
    vals_m = [year, 'Maison', '1 500 m', len(items_m), avg_m, avg_surf_m, round(avg_m * avg_surf_m), var_abs_m, var_pct_m]
    for c_i, v in enumerate(vals_m, 1):
        fmt = '#,##0 [$€]' if c_i in (5, 7) else None
        ce = cell(ws1, row, c_i, v, fill=fill_m, fmt=fmt)
        if c_i == 9 and isinstance(v, float):
            ce.font = Font(color='1a7a1a' if v > 0 else 'cc2200', bold=True, size=10)
    row += 1

for i, w in enumerate([8, 16, 10, 8, 15, 14, 20, 12, 10], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ==================== FEUILLE 2 : Maisons 800m vs 1500m ====================
ws2 = wb.create_sheet('Maisons 800m vs 1500m')

ws2.merge_cells('A1:G1')
c = ws2['A1']
c.value = 'Maisons — Comparaison rayon 800 m (inner) vs 1 500 m (total incluant 800m)'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='0d2137')
c.alignment = center
ws2.row_dimensions[1].height = 24
ws2.row_dimensions[2].height = 6

hdrs2 = ['Année', 'Nb tx 800m', 'Prix/m² 800m', 'Nb tx 1500m', 'Prix/m² 1500m', 'Tx anneau 800-1500m', 'Prix/m² anneau']
for c_i, h in enumerate(hdrs2, 1):
    cell(ws2, 3, c_i, h, fill=hdr2_fill, fnt=bold_w)
ws2.row_dimensions[3].height = 22

for ri, year in enumerate(years, 4):
    i8 = maison_data['r800'][str(year)]
    i15 = maison_data['r1500'][str(year)]
    i_anneau = [i for i in i15 if i['dist'] > 800]
    avg8  = round(sum(i['ppm2'] for i in i8) / len(i8))  if i8  else 0
    avg15 = round(sum(i['ppm2'] for i in i15) / len(i15)) if i15 else 0
    avg_an = round(sum(i['ppm2'] for i in i_anneau) / len(i_anneau)) if i_anneau else 0
    fill = alt3_fill if ri % 2 == 0 else alt4_fill
    vals = [year, len(i8), avg8, len(i15), avg15, len(i_anneau), avg_an]
    for c_i, v in enumerate(vals, 1):
        fmt = '#,##0 [$€]' if c_i in (3, 5, 7) else None
        cell(ws2, ri, c_i, v, fill=fill, fmt=fmt)

for i, w in enumerate([8, 14, 14, 14, 14, 22, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ==================== FEUILLE 3 : Transactions détaillées ====================
ws3 = wb.create_sheet('Transactions détaillées')
hdrs3 = ['Année', 'Mois', 'Type', 'Rayon', 'Prix (€)', 'Surface (m²)', 'Prix/m²', 'Dist. centre (m)']
for c_i, h in enumerate(hdrs3, 1):
    cell(ws3, 1, c_i, h, fill=hdr_fill, fnt=bold_w)
ws3.row_dimensions[1].height = 20

row3 = 2
for year in years:
    for item in sorted(all_data[str(year)]['Appartement'], key=lambda x: x['mois']):
        fill = alt1_fill if row3 % 2 == 0 else alt2_fill
        vals = [year, item['mois'], 'Appartement', '800m', item['prix'], item['surface'], item['ppm2'], '']
        for c_i, v in enumerate(vals, 1):
            cell(ws3, row3, c_i, v, fill=fill, fmt='#,##0' if c_i in (5, 7) else None)
        row3 += 1
    for item in sorted(maison_data['r1500'][str(year)], key=lambda x: x['mois']):
        rayon = '800m' if item['dist'] <= 800 else '1500m'
        fill = alt3_fill if row3 % 2 == 0 else alt4_fill
        vals = [year, item['mois'], 'Maison', rayon, item['prix'], item['surface'], item['ppm2'], item['dist']]
        for c_i, v in enumerate(vals, 1):
            cell(ws3, row3, c_i, v, fill=fill, fmt='#,##0' if c_i in (5, 7) else None)
        row3 += 1

for i, w in enumerate([8, 10, 14, 8, 14, 14, 12, 16], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

path = r'C:\Users\vince\Desktop\Projets IA Antigravity\Afficheur_Tram_Nantes\DVF_Felix_Faure.xlsx'
wb.save(path)
print('Saved:', path)
